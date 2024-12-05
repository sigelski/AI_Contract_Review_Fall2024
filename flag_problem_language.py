import re
import openpyxl
import nltk
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.corpus import stopwords
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# Ensure NLTK resources are downloaded
nltk.download('punkt', quiet=True)
nltk.download('stopwords', quiet=True)

# Constants
AUBURN_TITLES = ["Auburn's Preferred Language", "Auburn Preferred Language"]
EXCLUDED_SHEETS = {"INDEX", "template", "CONTACTS"}
STOP_WORDS = set(stopwords.words('english'))
SIMILARITY_THRESHOLD = 0.20
CONTRACT_INPUT_PATH = 'contract_to_txt.txt'
FLAGGED_OUTPUT_PATH = 'flagged_contract_to_txt.txt'

def extract_sheet_data(sheet):
    """
    Extracts relevant data from an Excel sheet.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel sheet to extract data from.

    Returns:
        dict: A dictionary containing extracted data.
    """
    data = {
        "Auburn's Preferred Language": [],
        'Common Problems': [],
        'Why': [],
        '1st response to Sponsor': []
    }

    current_key = None
    for row in sheet.iter_rows(values_only=True):
        if not row[1]:
            continue  # Skip empty rows in the second column

        cell_value = row[1]
        if cell_value in AUBURN_TITLES:
            current_key = "Auburn's Preferred Language"
            continue

        if cell_value in data:
            current_key = cell_value
            continue

        if current_key == 'Common Problems':
            data['Common Problems'].append(cell_value)
            data['Why'].append(row[2] if len(row) > 2 else None)
            data['1st response to Sponsor'].append(row[3] if len(row) > 3 else None)
        elif current_key == "Auburn's Preferred Language":
            data[current_key].append(cell_value)

    return data

def pretty_print_nested_values(dictionary):
    """
    Generates a readable string representation of a nested dictionary.

    Args:
        dictionary (dict): The nested dictionary to print.

    Returns:
        str: A formatted string representing the dictionary.
    """
    output = []
    for page_title, inner_dict in dictionary.items():
        output.append(page_title)
        for cell_title, values in inner_dict.items():
            output.append(f"\t{cell_title}:")
            for value in values:
                output.append(f"\t\t{value if value is not None else 'None'}")
        output.append('')  # Add a blank line for separation
    return "\n".join(output)

def jaccard_similarity(sentence1, sentence2):
    """
    Calculates the Jaccard similarity between two sentences.

    Args:
        sentence1 (str): The first sentence.
        sentence2 (str): The second sentence.

    Returns:
        float: Jaccard similarity score.
    """
    words1 = set(word_tokenize(sentence1.lower())) - STOP_WORDS
    words2 = set(word_tokenize(sentence2.lower())) - STOP_WORDS
    intersection = words1.intersection(words2)
    union = words1.union(words2)
    return len(intersection) / len(union) if union else 0.0

def cosine_similarity_score(sen1, sen2):
    """
    Calculates the cosine similarity between two sentences using TF-IDF vectors.

    Args:
        sen1 (str): The first sentence.
        sen2 (str): The second sentence.

    Returns:
        float: Cosine similarity score.
    """
    vectorizer = TfidfVectorizer(min_df=1, stop_words='english')
    vectors = vectorizer.fit_transform([sen1, sen2])
    sim_matrix = cosine_similarity(vectors)
    return sim_matrix[0][1]

def flag_sentences(tnc_dictionary):
    """
    Flags sentences in the contract text based on similarity to common problems.

    Args:
        tnc_dictionary (dict): The dictionary containing terms and conditions data.

    Returns:
        dict: A dictionary of flagged sentences with relevant metadata.
    """
    with open(CONTRACT_INPUT_PATH, 'r', encoding='utf-8') as file:
        contract_text = file.read()

    contract_sentences = sent_tokenize(contract_text)
    flagged_sentences = {}

    for sheet_name, sub_dict in tnc_dictionary.items():
        common_problems = sub_dict.get("Common Problems", [])
        preferred_language = sub_dict.get("Auburn's Preferred Language", [])
        why = sub_dict.get("Why", [])
        response = sub_dict.get("1st response to Sponsor", [])

        for idx, sentence in enumerate(contract_sentences):
            for i, problem in enumerate(common_problems):
                similarity = cosine_similarity_score(problem, sentence)
                if similarity > SIMILARITY_THRESHOLD:
                    flagged_sentences[sentence] = {
                        "Problem Category": sheet_name,
                        "Common Problems": problem,
                        "Preferred Language": preferred_language,
                        "Why": why[i] if i < len(why) else None,
                        "1st response to Sponsor": response[i] if i < len(response) else None,
                        "Confidence": similarity
                    }

    return flagged_sentences

def flag_problem_language(tnc_path_in):
    """
    Main function to flag problematic language in the contract based on T&Cs.

    Args:
        tnc_path_in (str): Path to the Excel file containing T&Cs.

    Outputs:
        Writes the flagged contract to a text file.
    """
    try:
        workbook = openpyxl.load_workbook(tnc_path_in)
    except FileNotFoundError:
        print(f"Error: The file {tnc_path_in} was not found.")
        return
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Error: The file {tnc_path_in} is not a valid Excel file.")
        return

    sheet_names = workbook.sheetnames
    tnc_dictionary = {}

    for sheet_name in sheet_names:
        if sheet_name not in EXCLUDED_SHEETS:
            sheet = workbook[sheet_name]
            tnc_dictionary[sheet_name] = extract_sheet_data(sheet)

    flagged_sentences = flag_sentences(tnc_dictionary)

    try:
        with open(CONTRACT_INPUT_PATH, 'r', encoding='utf-8') as file:
            contract_lines = file.read()
    except FileNotFoundError:
        print(f"Error: The file {CONTRACT_INPUT_PATH} was not found.")
        return

    tokenized_lines = sent_tokenize(contract_lines)
    modified_lines = []

    for line in tokenized_lines:
        stripped_line = line.strip()
        if stripped_line in flagged_sentences:
            details = flagged_sentences[stripped_line]
            modified_lines.append("\t\t" + "*" * 91)
            modified_lines.append("\t[POTENTIAL PROBLEMATIC LANGUAGE DETECTED]")
            modified_lines.append(line)
            for key, value in details.items():
                modified_lines.append(f"\t\t{key}: {value}")
            modified_lines.append("\t\t[END POTENTIAL PROBLEMATIC LANGUAGE]")
            modified_lines.append("*" * 91)
        else:
            modified_lines.append(line)

    try:
        with open(FLAGGED_OUTPUT_PATH, 'w', encoding='utf-8') as f:
            f.write('\n'.join(modified_lines))
        print(f"Flagged contract has been saved to {FLAGGED_OUTPUT_PATH}.")
    except IOError as e:
        print(f"Error writing to file {FLAGGED_OUTPUT_PATH}: {e}")

if __name__ == "__main__":
    pass 